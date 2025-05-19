
import pandas as pd
import ast  # Per convertire stringhe in liste Python
import numpy as np

import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.views import SheetView
from openpyxl.worksheet.views import Pane

def calcola_valore_migliore(df, colonna_esecuzione, colonna_valore, colonna_risultato):
    """
    Funzione generica per calcolare il valore migliore per una colonna.
    
    Parametri:
    - df: DataFrame principale
    - colonna_esecuzione: nome della colonna di raggruppamento (es. "Esecuzione")
    - colonna_valore: nome della colonna di cui trovare il valore minimo (es. "Loss Finale")
    - colonna_risultato: nome della colonna in cui inserire il valore migliore (es. "Loss Migliore")
    
    Ritorna:
    - Il DataFrame con le righe che hanno valori migliori non-null
    """
    # Filtra solo le colonne rilevanti
    df_filtrato = df[[colonna_esecuzione, colonna_valore]]
    
    # Trova il valore minimo per ogni gruppo
    df_min = df_filtrato.groupby(colonna_esecuzione)[colonna_valore].min().reset_index()
    
    # Inizializza la colonna risultato con None
    df[colonna_risultato] = None
    
    # Per ogni riga del DataFrame con i minimi, trova la corrispondente riga nel DataFrame originale
    for _, row in df_min.iterrows():
        esecuzione = row[colonna_esecuzione]
        valore_minimo = row[colonna_valore]
        
        # Trova le righe che corrispondono
        matches = df[(df[colonna_esecuzione] == esecuzione) & (df[colonna_valore] == valore_minimo)]
        
        if not matches.empty:
            # Prendi l'indice della prima occorrenza
            idx_min = matches.index[0]
            
            # Assegna il valore minimo alla colonna risultato
            df.at[idx_min, colonna_risultato] = valore_minimo
        else:
            print(f"Nessun match trovato per {colonna_esecuzione}: {esecuzione}, {colonna_valore}: {valore_minimo}")
    
    # Ritorna le righe con valori migliori non-null
    return df[df[colonna_risultato].notna()]

def Analisi(file_excel,foglio):
    # Crea un dataframe vuoto
    results_df = pd.DataFrame()
    # Carica dati esistenti o crea nuovo dataframe
    if os.path.exists(file_excel):
        try:
            wb = load_workbook(file_excel)
            if foglio in wb.sheetnames:
                ws = wb[foglio]
                existing_data = []
                for row in ws.iter_rows(values_only=False):
                    row_data = []
                    for cell in row:
                        if cell.data_type == "f":
                            row_data.append(cell.value)
                        else:
                            row_data.append(cell.value)
                    existing_data.append(row_data)
                
                if existing_data:
                    header = existing_data[0]
                    dati_esistenti = pd.DataFrame(existing_data[1:], columns=header)
                else:
                    dati_esistenti = pd.DataFrame(columns=results_df.columns)
            else:
                dati_esistenti = pd.DataFrame(columns=results_df.columns)
        except Exception as e:
            print(f"Errore durante la lettura: {e}")
            dati_esistenti = pd.DataFrame(columns=results_df.columns)
    else:
        dati_esistenti = pd.DataFrame(columns=results_df.columns)


    # Assicurati di aver creato la colonna "Esecuzione originale" 
    dati_esistenti["Esecuzione originale"] = dati_esistenti["Esecuzione"].copy()

    # Definisci le metriche da calcolare
    metriche = [
        {"valore": "Loss Finale", "migliore": "Loss Migliore per ogni esecuzione"},
        {"valore": "Gradiente Finale", "migliore": "Gradiente Migliore per ogni esecuzione"},
        {"valore": "Tempo di Esecuzione (s)", "migliore": "Tempo Migliore per ogni esecuzione"}
    ]

    # Inizializza il DataFrame che conterrà tutti i valori migliori
    dati_migliori = pd.DataFrame()

    # Ciclo sulle metriche
    for metrica in metriche:
        # Calcola i valori migliori per questa metrica
        df_metrica = calcola_valore_migliore(
            dati_esistenti, 
            "Esecuzione", 
            metrica["valore"], 
            metrica["migliore"]
        )
        
        # Aggiungi i risultati al DataFrame principale
        dati_migliori = pd.concat([dati_migliori, df_metrica], ignore_index=True)

    # Ordina i dati per la colonna di esecuzione originale
    dati_migliori = dati_migliori.sort_values(by='Esecuzione originale')
    
    return dati_migliori
    
