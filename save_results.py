import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.worksheet.views import SheetView, Pane
from openpyxl.styles import Border, Side

def applica_bordi(ws):
    """Applica un bordo nero a tutte le celle contenenti dati nel foglio di lavoro."""
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border

def adjust_column_width(ws):
    """Regola automaticamente la larghezza delle colonne"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
        
        # Aggiunge il wrap text per le celle lunghe
        for cell in column:
            cell.alignment = Alignment(wrap_text=True)

def salva_matrice(df,nome_df, file_excel):
    
    # Determina numero di esecuzione
    if os.path.exists(file_excel):
        try:
            risultati = pd.read_excel(file_excel, sheet_name="Risultati")
            numero_esecuzione = risultati["Esecuzione"].max() 
        except:
            numero_esecuzione = 1
    else:
        numero_esecuzione = 1
    
    # Salva in nuovo sheet
    sheet_name = f"Dati_{nome_df}"
    
    #Salva il Dataframe

    if os.path.exists(file_excel):
        # File esiste: usa append con gestione sheet esistente
        with pd.ExcelWriter(file_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.head().to_excel(writer, sheet_name=sheet_name)
    else:
        # File non esiste: crea nuovo file
        with pd.ExcelWriter(file_excel, engine='openpyxl', mode='w') as writer:
            df.head().to_excel(writer, sheet_name=sheet_name)
    
    # Dopo aver salvato i dati, regola le colonne
    if os.path.exists(file_excel):
        wb = load_workbook(file_excel)
        ws = wb[sheet_name]
        adjust_column_width(ws)
        wb.save(file_excel)
    with open("output.txt", "a") as f:
        print(f"\nMatrice X salvata in {sheet_name} di {file_excel}",file=f)



def save_results_to_excel(results_df, file_excel,foglio):
    """
    Salva il dataframe dei risultati in un file Excel con formattazione.
    
    Args:
        results_df: Dataframe pandas con i risultati
        file_excel: Nome del file Excel di output
    """
    # Carica dati esistenti o crea nuovo dataframe
    if os.path.exists(file_excel):
        try:
            wb = load_workbook(file_excel)
            if foglio in wb.sheetnames:
                ws = wb[foglio]
                existing_data = []
                for row in ws.iter_rows(values_only=False):
                    row_data = []
                    for cell in row:
                        if cell.data_type == "f":
                            row_data.append(cell.value)
                        else:
                            row_data.append(cell.value)
                    existing_data.append(row_data)
                
                if existing_data:
                    header = existing_data[0]
                    dati_esistenti = pd.DataFrame(existing_data[1:], columns=header)
                else:
                    dati_esistenti = pd.DataFrame(columns=results_df.columns)
            else:
                dati_esistenti = pd.DataFrame(columns=results_df.columns)
        except Exception as e:
            print(f"Errore durante la lettura: {e}")
            dati_esistenti = pd.DataFrame(columns=results_df.columns)
    else:
        dati_esistenti = pd.DataFrame(columns=results_df.columns)

    # Gestione numero esecuzione
    if not dati_esistenti.empty:
        nuovo_numero_esecuzione = dati_esistenti["Esecuzione"].max() 
        if pd.isna(nuovo_numero_esecuzione):
            nuovo_numero_esecuzione = 0
        if (len(dati_esistenti) % 5 == 0) :
            nuovo_numero_esecuzione += 1
    else:
        nuovo_numero_esecuzione = 1

    if foglio=="Analisi":
        results_df["Esecuzione"] =results_df["Esecuzione originale"].copy()
        results_df.drop(["Esecuzione originale"], axis = 1)
    else:
        results_df["Esecuzione"] = nuovo_numero_esecuzione

    # Unisci i dati
    df_finale = pd.concat([
        dati_esistenti[results_df.columns],
        results_df[results_df.columns]
    ], ignore_index=True)

    # Salva il DataFrame nel file Excel
    if not os.path.exists(file_excel):
        with pd.ExcelWriter(file_excel, engine='openpyxl', mode='w') as writer:
            df_finale.to_excel(writer, sheet_name=foglio, index=False)
    else:
        with pd.ExcelWriter(file_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_finale.to_excel(writer, sheet_name=foglio, index=False)

    # Formattazione
    wb = load_workbook(file_excel)
    ws = wb[foglio]
    
    colors = ['FFD3D3D3', 'FFA5A5A5', 'FFE6E6E6', 'FFC9C9C9']
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        esecuzione = ws.cell(row=idx, column=1).value
        color_index = (esecuzione - 1) % len(colors)
        fill = PatternFill(start_color=colors[color_index], end_color=colors[color_index], fill_type="solid")
        
        for cell in row:
            cell.fill = fill
            if cell.column == 8  :
                # Controlla il valore della cella e la sua formula
                formula = cell.value
                # Se la formula non contiene già "Vedi grafico", solo allora modifica
                if not (isinstance(formula, str) and "Vedi grafico" in formula):
                    original_value = cell.value  # Salva il valore originale
                    if original_value:  # Assicurati che il valore non sia None o vuoto
                        
                        cell.value = f'=HYPERLINK("{original_value}", "Vedi grafico")'

    applica_bordi(ws)

    adjust_column_width(ws)
    
    # Blocca la prima riga e le prime due colonne
    ws.freeze_panes = 'D2'

    wb.save(file_excel)
    print(f"\n📂 Risultati aggiornati nel sheet 'Risultati' di {file_excel}")



