
import pandas as pd
import ast  # Per convertire stringhe in liste Python
import numpy as np

import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.views import SheetView
from openpyxl.worksheet.views import Pane
from openpyxl.styles import Border, Side

def Analisi(file_excel, foglio):
    wb = load_workbook(file_excel)
    ws = wb[foglio]
    data = ws.values
    cols = next(data)
    df = pd.DataFrame(data, columns=cols)

    df["Esecuzione originale"] = df["Esecuzione"].copy()

    for col in ["Loss Finale", "Gradiente Finale", "Tempo di Esecuzione (s)"]:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    idx_best_loss = df.groupby("Esecuzione")["Loss Finale"].idxmin()
    idx_best_time = df.groupby("Esecuzione")["Tempo di Esecuzione (s)"].idxmin()

    analisi_globale = df.loc[sorted(set(idx_best_loss).union(set(idx_best_time))), ["Metodo", "Loss Finale", "Tempo di Esecuzione (s)"]]

    if "Analisi" in wb.sheetnames:
        del wb["Analisi"]
    ws_new = wb.create_sheet("Analisi")

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    colors = ['FFD3D3D3', 'FFA5A5A5', 'FFE6E6E6', 'FFC9C9C9']

    # Scrivi header
    for c_idx, col in enumerate(df.columns, 1):
        cell = ws_new.cell(row=1, column=c_idx, value=col)
        cell.font = bold_font
        cell.border = thin_border

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        esecuzione = getattr(row, "Esecuzione")
        color_index = (int(esecuzione) - 1) % len(colors)
        row_fill = PatternFill(start_color=colors[color_index], end_color=colors[color_index], fill_type="solid")

        for c_idx, value in enumerate(row, 1):
            cell = ws_new.cell(row=r_idx, column=c_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    # Evidenzia migliori righe
    for idx in idx_best_loss:
        row_excel = idx + 2
        col_loss = df.columns.get_loc("Loss Finale") + 1
        for col in range(1, len(df.columns) + 1):
            ws_new.cell(row=row_excel, column=col).fill = green_fill
            ws_new.cell(row=row_excel, column=col).border = thin_border
        ws_new.cell(row=row_excel, column=col_loss).fill = yellow_fill

    for idx in idx_best_time:
        row_excel = idx + 2
        col_time = df.columns.get_loc("Tempo di Esecuzione (s)") + 1
        for col in range(1, len(df.columns) + 1):
            cell = ws_new.cell(row=row_excel, column=col)
            if cell.fill.start_color.rgb != yellow_fill.start_color.rgb:
                cell.fill = green_fill
            cell.border = thin_border
        ws_new.cell(row=row_excel, column=col_time).fill = yellow_fill

    # Larghezza colonne
    for column in ws_new.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        col_letter = get_column_letter(column[0].column)
        ws_new.column_dimensions[col_letter].width = (max_len + 2) * 1.2

    ws_new.freeze_panes = 'D2'

    # Analisi globale per metodo
    if "Analisi_globale" in wb.sheetnames:
        del wb["Analisi_globale"]
    ws_glob = wb.create_sheet("Analisi_globale")
    ws_glob.append(["Metodo","Esecuzione (Loss-Min)", "Numero di sottoblocchi (Loss-Min)", "Loss Minimo", "Esecuzione (Tempo-Min)","Numero di sottoblocchi (Tempo-Min)", "Tempo Minimo","Esecuzione (Loss-Max)","Numero di sottoblocchi (Loss-Max)", "Loss Max","Esecuzione (Tempo-Max)", "Numero di sottoblocchi (Tempo-Max)", "Tempo Max"])

    for metodo, group in df.groupby("Metodo"):
        # Minimo loss e indice relativo
        min_loss = group["Loss Finale"].min()
        row_loss = group[group["Loss Finale"] == min_loss].iloc[0]
        esec_loss = row_loss["Esecuzione"]
        num_block_loss = row_loss["Numero sottoblocchi"]

        # Max loss e indice relativo
        max_loss = group["Loss Finale"].max()
        row_loss_max = group[group["Loss Finale"] == max_loss].iloc[0]
        esec_loss_max = row_loss_max["Esecuzione"]
        num_block_loss_max = row_loss_max["Numero sottoblocchi"]

        # Minimo tempo e indice relativo
        min_time = group["Tempo di Esecuzione (s)"].min()
        row_time = group[group["Tempo di Esecuzione (s)"] == min_time].iloc[0]
        esec_time = row_time["Esecuzione"]
        num_block_time = row_time["Numero sottoblocchi"]

        # Max tempo e indice relativo
        max_time = group["Tempo di Esecuzione (s)"].max()
        row_time_max = group[group["Tempo di Esecuzione (s)"] == max_time].iloc[0]
        esec_time_max = row_time_max["Esecuzione"]
        num_block_time_max = row_time_max["Numero sottoblocchi"]

        ws_glob.append([metodo, esec_loss,num_block_loss, min_loss, esec_time,num_block_time, min_time,esec_loss_max,num_block_loss_max, max_loss, esec_time_max,num_block_time_max, max_time])


    # Formattazione Analisi_globale
    for row in ws_glob.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    for column in ws_glob.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        col_letter = get_column_letter(column[0].column)
        ws_glob.column_dimensions[col_letter].width = (max_len + 2) * 1.2

    # Blocca la prima riga e la prima colonna
    ws_glob.freeze_panes = 'B2'

    wb.save(file_excel)
    
    print("✅ Analisi salvata nei fogli 'Analisi' e 'Analisi_globale'")
    return None

